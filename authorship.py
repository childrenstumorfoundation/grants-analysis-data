
# the program takes the "researcher" cell in the ctf-grant excel, search for it in publication pdfs,
# and find all files that the author appears in
import openpyxl
import codecs
import regex as re
from openpyxl.utils import get_column_letter as c_l

# given the list of publications where the same author name appears,
# the program find the one with publication date closest to the grant end year,
# and narrow down the possible publication of a grant to one.

wb = openpyxl.load_workbook('CTF-Grant.xlsx')
sheet = wb.active
numRow = sheet.max_row

wb1 = openpyxl.load_workbook('Publication.xlsx')
sheet1 = wb1['selection']
numRow1 = sheet1.max_row

wb2 = openpyxl.load_workbook('author_grantyear_paperIDs.xlsx')
sheet2 = wb2.active
numRow2 = sheet2.max_row

name_not_appear = []
def searchByAuthorshipDate(list):
# build a dictionary to pair up the author - end year:
    dictionary = {}
# build on top of dictionary, add the grant number
    dictionary1 = {}
# build a list of author names in text format, without the duplicates.
    mylist=[]

    for rowindex in range (3,numRow+1):
        author = sheet['O'+str(rowindex)].value
        mylist.append(author)
        startYear = sheet['L'+str(rowindex)].value
        # use a nested list within a dictionary to store more than one value(grant start year) to the key(author),
#each value followed by its grant number
        grant_number = sheet['C'+str(rowindex)].value
        nested_list =[]
        nested_list.append(startYear)
        nested_list.append(grant_number)
        dictionary.setdefault(author, []).append(startYear)
        dictionary1.setdefault(author, []).append(nested_list)

# the set() built-in function get unique collection of author names.
    finallist = set(mylist)
    a = 2
# loop through all author names:
    for name in finallist:
        authorname = re.compile(str(name), re.I)
        # authorname = re.compile("(str(name)){e<2}", re.I)
        print(str(name)+':  '+ str(dictionary[name]))
        # text file name
        sheet2['A' + str(a)]=name
        # grant year
        sheet2['B'+str(a)]=str(dictionary[name])

        listTOdate = []
        for i in range(0, len(list)):
            file = codecs.open(str(list[i]), encoding='utf-8')
            # load it once
            filetext = file.read()

            if re.search(authorname, filetext):
                listTOdate.append(list[i])
        print("papers with the author's name")
        print(listTOdate)

        if len(listTOdate)==0:
            name_not_appear.append(name)
            # record grant ID1 for author without potential file list
            column2 = 9
            for i1 in range(0, len(dictionary1[name])):
                sheet2[c_l(column2) + str(a)] = dictionary1[name][i1][1]
                column2+=1

# 3, date approach
# get the publication year of the list of publication papers from publication.xlsx
        dictionary2 = {}
        # pair up grant year and paper names
        for publication in listTOdate:
            for rowindex1 in range (2,numRow1+1):
                if publication == sheet1['L'+str(rowindex1)].value:
                    publicationyear = sheet1['B' + str(rowindex1)].value
            list_grantyear = dictionary1[name]

            column = 3
            column1 = 9
            for i in range(0, len(list_grantyear)):
                if list_grantyear[i][0] <= publicationyear:
                    dictionary2.setdefault(list_grantyear[i][0],[]).append(publication)
                    if sheet2[c_l(column+i) + str(a)].value == None:
                        sheet2[c_l(column+i) + str(a)] = publication
                    else:
                        newvalue = str(sheet2[c_l(column+i) + str(a)].value) + ' , '+ str(publication)
                        sheet2[c_l(column + i) + str(a)]=newvalue
                # else:
                #     print('grant year:'+str(list_grantyear[i]) + '  publication year:'+str(publicationyear))
                #     print(str(publication)+' is not possible')
                sheet2[c_l(column1+i)+str(a)] =list_grantyear[i][1]
        a+=1
        print(dictionary2)

list = []
for row in range (2, numRow1):
    filename = sheet1['L'+str(row)].value
    list.append(filename)

searchByAuthorshipDate(list)
wb2.save('author_grantyear_paperIDs.xlsx')

print(len(name_not_appear))







