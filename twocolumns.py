# it writes the local storage and filename column to publication

import openpyxl
wb = openpyxl.load_workbook('Publication.xlsx')
sheet = wb['selection']
num_row = sheet.max_row

for i in range (2,num_row+1):
    doi = sheet['E'+str(i)].value
    filename = str(doi).replace('/','_') + '.pdf'
    sheet['J'+str(i)] = filename

    filename = sheet['J' + str(i)].value
    pdfloc = "grants-analysis-data\\papers\\pdf\\" + str(filename)
    sheet['I' + str(i)] = pdfloc

    textname = str(filename).replace('.pdf', '.txt')
    textloc = "grants-analysis-data\\papers\\txt\\" + str(textname)
    sheet['K' + str(i)] = textloc

    sheet['L' + str(i)] = textname

# add the column of "paper abstract" to 'publication excel'

sheet1 = wb['My Library']
num_row1 = sheet1.max_row

for i in range(2, num_row1 + 1):
    my_doi = sheet1['I' + str(i)].value
    for i1 in range(2, num_row + 1):
        doi = sheet['E' + str(i1)].value
        if my_doi == doi:
            sheet['M' + str(i1)] = sheet1['K' + str(i)].value

wb.save('Publication.xlsx')

