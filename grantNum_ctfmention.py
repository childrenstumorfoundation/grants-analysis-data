# the program first searches for the grant number in all files in function "searchByGrantNum"
# then search for "ctf" string in the list of files with grant number, in function "searchByMention"
# finally, enter the confirmed files' paperID & grantID

import regex as re
import codecs
import openpyxl

list_with_grantnum = []
dictionary = {}
nomentionList = []

wb1 = openpyxl.load_workbook('Publication.xlsx')
sheet1 = wb1['selection']
numRow1 = sheet1.max_row

# searches for the grant number string in all files
def searchByGrantNum(mylist):

    grantRegex = re.compile(r'\d{4}[A-Z]-\d{2}-\d{3}')
    grantRegex1 = re.compile(r'\d{4}-\d{2}-\d{3}[A-Z]')
    grantRegex2 = re.compile(r'\d{4}-\d{2}-[A-Z]')
    grantRegex3 = re.compile(r'\d{4}-\d{2}-\d{3}')

    for i in range(0, len(mylist)):
        list = []
        filename = mylist[i]
        file = codecs.open(str(filename), encoding = 'utf-8')
        # load it once
        filetext = file.read()
        if re.search(grantRegex,filetext) or re.search(grantRegex1,filetext) or re.search(grantRegex2,filetext) or re.search(grantRegex3,filetext):
            list_with_grantnum.append(filename)

            list1= re.findall(grantRegex,filetext)
            list2=re.findall(grantRegex1,filetext)
            list3=re.findall(grantRegex2, filetext)
            list4 = re.findall(grantRegex3,filetext)

            if list2!=[]:
                list = list1+list2+list3
            else:
                list = list1+list3+list4
            print(list)
            dictionary[filename]=list

        else:
            print('no')
    return dictionary
    return list_with_grantnum


# loop through text files of papers with grant number, and search for "CTF" string for double verification.
def searchByMention(list):

    ctf = re.compile("(Children's Tumor Foundation){e<=4}",flags=re.IGNORECASE)
    ctf1 = re.compile("ctf",re.I)
    nnf = re.compile("(National Neurofibromatosis Foundation){e<=3}", flags=re.IGNORECASE)
    nnf1 = re.compile("nnf", re.I)

    for i in range(0, len(list)):
        filename = list[i]
        print(str(filename))
        textname = str(filename).replace('.pdf', '.txt')
        file = codecs.open(str(textname), encoding = 'utf-8')
        # load it once
        filetext = file.read()

        if re.search(ctf, filetext) or re.search(ctf1, filetext) or re.search(nnf, filetext) or re.search(nnf1, filetext):
            print("YES")

        else:
            print('NO')
            nomentionList.append(filename)
    return nomentionList

orignal_list = []
for row in range (2, numRow1):
    filename = sheet1['L'+str(row)].value
    orignal_list.append(filename)

searchByGrantNum(orignal_list)
searchByMention(list_with_grantnum)

print('Following is the list of files with grant number:')
print(list_with_grantnum)
print(len(list_with_grantnum))

print('Following is the list of files without ctf mentions:')
print(nomentionList)
print(len(nomentionList))

print('Following is the list of files that are confirmed to be funded by CTF')
confirmed_list = [x for x in list_with_grantnum if x not in nomentionList]
print(confirmed_list)
print(len(confirmed_list))

print(dictionary)


#enter paperID and grantID into a separate excel

wb = openpyxl.load_workbook('Workbook1.xlsx')
sheet = wb['Sheet1']

for i in range(0, len(confirmed_list)):
    sheet['A' + str(i + 2)] = confirmed_list[i]
    string = str(confirmed_list[i]).replace('_', '/')
    doi_string = string.split('.pdf')[0]
    sheet['C' + str(i + 2)] = doi_string
    sheet['B' + str(i + 2)] = str(dictionary[confirmed_list[i]])

wb.save('Workbook1.xlsx')