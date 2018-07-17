# print the grant info: grant number 'C'
#  title 'D', start date 'K', end date 'M', Researcher 'O',research organization 'P', abstract'F'
# print the paper info: paperID 'E'
# title 'D', pubication date 'G', author 'C', abstract 'M', url 'F'

# return Y: add to dictionary{}, grantnumber-paperID pair
# return N: delete it from "possible_list" 'D', given grant number
# return Maybe: do nothing

import openpyxl

wb = openpyxl.load_workbook('CTF-Grant.xlsx')
sheet = wb.active
numRow = sheet.max_row

wb1 = openpyxl.load_workbook('Publication.xlsx')
sheet1 = wb1['selection']
numRow1 = sheet1.max_row

wb2 = openpyxl.load_workbook('possible_list.xlsx')
sheet2 = wb2.active
numRow2 = sheet2.max_row

wb3 = openpyxl.load_workbook('list.xlsx')
sheet3 = wb3.active
numRow3 = sheet3.max_row

dictionary = {}
a=2

for i in range(3, numRow1+1):
    grant_num =sheet['C'+str(i)].value
    title=sheet['D'+str(i)].value
    start_date = sheet['K'+str(i)].value
    end_date = sheet['M'+str(i)].value
    researcher = sheet['O'+str(i)].value
    research_org=sheet['P'+str(i)].value
    grant_abstract=sheet['F'+str(i)].value

    for i1 in range (2, numRow2+1):
        if grant_num==sheet2['A'+str(i1)].value:
            string_of_dois=sheet2['D'+str(i1)].value
            if string_of_dois!=None:
                new_dois = string_of_dois.replace('_','/').replace('.txt','')
                dois=new_dois.split(',')
                num_remaining = len(dois)-1
                for doi in dois:
                    for i2 in range(2, numRow1+1):
                        if doi==sheet1['E'+str(i2)].value:
                            my_title = sheet1['D'+str(i2)].value
                            publication_date = sheet1['G'+str(i2)].value
                            my_author = sheet1['C'+str(i2)].value
                            my_abstract = sheet1['M'+str(i2)].value
                            my_url = sheet1['F'+str(i2)].value
                    print('''
       Grant Number: {} 
           Title: {}
           Start date: {}  |  End date: {}
           Researcher: {}  |  Research Organization: {}
           Grant Abstract:{} '''.format(grant_num, title, start_date, end_date, researcher, research_org,
                                        grant_abstract))
                    print('''
        Number of papers remaining: {}
        
        Paper ID: {}
            Title: {}
            Publication date: {}
            Author: {}
            Abstract: {}
            URL: {}
                '''.format(num_remaining,doi,my_title ,publication_date, my_author, my_abstract,my_url ))

                    check = True
                    while check:
                        userinput = input("Type in 'y'/ 'n' /'m' for 'YES /NO /MAYBE ': ")
                        if userinput=='y':
                            check1 = True
                        # ask for confirmation:
                            while check1:
                                second_input=input('Are you sure? y/n :')
                                if second_input=='y':
                            # enter the manually paired up grant num and paper id into "list" excel
                                    sheet3['A' + str(a)] = grant_num
                                    sheet3['B' + str(a)] = str(doi)
                                    a+=1
                                    wb3.save('list.xlsx')

                                    dictionary[grant_num] = doi
                                    print("The list of connections you've made:")
                                    print(dictionary)

                                    search = str(doi).replace('/','_')+'.txt'
                                    for row in range(2, numRow2+1):
                                        string = sheet2['D'+str(row)].value
                                        if string!= None:
                                            list1 = string.split(',')
                                            for item1 in list1:
                                                if item1==search:
                                                    list1.remove(item1)
                                            list2 = ','.join(list1)
                                            sheet2['D'+str(row)]=list2

                                    wb2.save('possible_list.xlsx')
                                    check = False
                                    check1 = False
                                elif second_input=='n':
                                    check1 = False
                                    check = True
                                else:
                                    print('wrong command')
                                    check1 = True

                        elif userinput=='n':
                            # remove it from the string of the list of paper IDs
                            dois.remove(doi)
                            myString = ''
                            dois1 = []
                            for item in dois:
                                item = str(item).replace('/', '_') + '.txt'
                                dois1.append(item)
                            sheet2['D' + str(i1)] = ','.join(dois1)
                            wb2.save('possible_list.xlsx')
                            check = False

                        elif userinput=='m':
                            check = False
                        else:
                            print('wrong command')
                    # print out the number of the remaining possible papers
                    num_remaining-=1

print(dictionary)