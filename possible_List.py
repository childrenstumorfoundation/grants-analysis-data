# this program takes the 'author_grantyear_paperID', rearrange it with grantID as index

import openpyxl
from openpyxl.utils import get_column_letter as c_l

wb = openpyxl.load_workbook('author_grantyear_paperIDs.xlsx')
sheet = wb.active
num_row = sheet.max_row

wb1 = openpyxl.load_workbook('possible_list.xlsx')
sheet1 = wb1.active

num=2
for i in range(2, num_row+2):
    grantyear=sheet['B'+str(i)].value
    author =sheet['A'+str(i)].value
    if grantyear==None:
        sheet1['B'+str(i)]=author
        num+=1
    else:
        list =grantyear.split(',')
        count = len(list)
        col = 9
        col1 = 3
        for year in list:
            sheet1['C'+str(num)]=year
            sheet1['B' + str(num)] =author
            sheet1['A' + str(num)] = sheet[c_l(col)+str(i)].value
            sheet1['D' + str(num)] = sheet[c_l(col1) + str(i)].value
            num+=1
            col+=1
            col1+=1
num_row1 = sheet1.max_row
for int in range (2, num_row1+1):
    string = sheet1['C'+str(int)].value
    new_string = string.replace('[','').replace(']','')
    sheet1['C'+str(int)]=new_string

wb1.save('possible_list.xlsx')